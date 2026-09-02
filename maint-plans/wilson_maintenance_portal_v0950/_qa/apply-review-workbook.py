#!/usr/bin/env python3
"""
Read a filled-in Wilson-Protocol-Review.xlsx and report what it changes.

This is the other half of the round trip. Without it the workbook is a promise:
"send it back and I will apply it" is worth nothing if applying it is a manual
transcription of 149 rows.

DELIBERATELY REPORT-ONLY BY DEFAULT. It prints the diff -- every score the owner
moved, every check whose scoring he flipped, every note he left -- and writes
nothing. `--apply` then edits assets/plan-config.js in place, matching each
option by its LABEL inside its own option set, and refuses if anything is
ambiguous. A script that silently half-applies a config change is worse than one
that makes you paste.

Usage:
    python3 apply-review-workbook.py Wilson-Protocol-Review.xlsx
    python3 apply-review-workbook.py Wilson-Protocol-Review.xlsx --apply
"""
from __future__ import annotations

import argparse
import json
import pathlib
import re
import subprocess
import sys

from openpyxl import load_workbook

HERE = pathlib.Path(__file__).parent
CONFIG = pathlib.Path("/home/claude/work/extracted/wilson_maintenance_portal_v0916/assets/plan-config.js")
EXAMPLE_MARK = "EXAMPLE ROW"

PROTOCOL_KEYS = {
    "Generic fallback": "generic", "Refrigeration": "refrigerator",
    "Dishwasher": "dishwasher", "Range, cooktop & oven": "cooking",
    "Icemaker (IMUC)": "icemaker", "Washer": "washer",
    "Laundry centre": "laundry", "Dryer": "dryer",
    "Vent hood": "ventilation", "Microwave & speed oven": "microwave",
    "Outdoor grill": "outdoor_grill", "Cooling - split system": "hvac_cooling",
    "Heat pump": "hvac_heatpump", "Furnace": "hvac_furnace",
    "Mini-split / ductless": "hvac_minisplit",
}


def live_config():
    """The protocol as it ships, straight from plan-config.js."""
    out = subprocess.run(
        ["node", "-e", """
        const fs=require('fs');global.window={};
        global.document={getElementById:()=>null,createElement:()=>({}),
          body:{classList:{add(){},remove(){}},appendChild(){}},
          querySelectorAll:()=>[],querySelector:()=>null};
        eval(fs.readFileSync(process.argv[1],'utf8'));
        const c=window.WILSON_CONFIG,A=window.WILSON_ANSWERS,out={};
        Object.keys(c.checkpointSets).forEach(function(k){
          out[k]={};
          (c.checkpointSets[k]||[]).forEach(function(cp){
            const a=A.for(k,cp.id);
            out[k][cp.name]={
              id:cp.id, kind:a.kind, scores:a.scores, optionSet:a.optionSet||null,
              options:(a.options||[]).map(function(o){
                return {code:o.code,label:o.label,
                        score:(Object.prototype.hasOwnProperty.call(o,'score')?o.score:undefined)};
              })
            };
          });
        });
        process.stdout.write(JSON.stringify(out));
        """, str(CONFIG)], capture_output=True, text=True, check=True)
    return json.loads(out.stdout)


def parse_score(raw):
    """A cell value -> a score, None for 'scores nothing', or 'drop'."""
    text = str(raw if raw is not None else "").strip().lower()
    if text in ("", "-"):
        return "blank"
    if text in ("none", "no score", "nothing", "n/a", "na"):
        return None
    if text == "drop":
        return "drop"
    m = re.match(r"^([1-5])(\s*of\s*5)?$", text)
    if m:
        return int(m.group(1))
    return ("bad", text)


def read_workbook(path):
    wb = load_workbook(path, data_only=True)
    checks, answers, notes = [], [], []

    if "Should it score" in wb.sheetnames:
        ws = wb["Should it score"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or EXAMPLE_MARK in str(row[0]):
                continue
            proto, name, _prompt, _how, today, verdict, note = (list(row) + [None] * 7)[:7]
            if verdict or note:
                checks.append({"protocol": proto, "check": name, "today": today,
                               "verdict": str(verdict).strip() if verdict else "",
                               "note": str(note).strip() if note else ""})

    if "What it's worth" in wb.sheetnames:
        ws = wb["What it's worth"]
        proto = check = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cells = (list(row) + [None] * 7)[:7]
            if cells[0] and EXAMPLE_MARK in str(cells[0]):
                continue
            # The appliance and check names are written once per group.
            if cells[0]:
                proto = str(cells[0]).strip()
            if cells[1]:
                check = str(cells[1]).strip()
            label, meaning, draft, mine, note = cells[2], cells[3], cells[4], cells[5], cells[6]
            if not label:
                continue
            if mine is None and not note:
                continue
            answers.append({"protocol": proto, "check": check, "label": str(label).strip(),
                            "meaning": meaning, "draft": draft,
                            "value": parse_score(mine),
                            "raw": mine,
                            "note": str(note).strip() if note else ""})
    return checks, answers, notes


def draft_number(draft):
    text = str(draft or "").strip().lower()
    if text in ("none", ""):
        return None
    if text in ("n/a", "na"):
        return "n/a"
    m = re.match(r"^([1-5])", text)
    return int(m.group(1)) if m else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--apply", action="store_true",
                    help="edit plan-config.js in place (default is report only)")
    args = ap.parse_args()

    cfg = live_config()
    checks, answers, _ = read_workbook(args.workbook)

    print("=" * 74)
    print("WHAT THIS WORKBOOK CHANGES")
    print("=" * 74)

    problems = []

    # ---- which checks should score -----------------------------------------
    if checks:
        print("\nSCORING DECISIONS (%d row%s answered)" % (len(checks), "" if len(checks) == 1 else "s"))
        for row in checks:
            key = PROTOCOL_KEYS.get(str(row["protocol"]).strip())
            live = (cfg.get(key) or {}).get(str(row["check"]).strip()) if key else None
            if not live:
                problems.append("cannot find check %r under %r" % (row["check"], row["protocol"]))
                continue
            now = "Yes" if live["scores"] else "No"
            want = row["verdict"]
            flag = "  <-- CHANGE" if want and want.lower() not in ("", now.lower(), "not sure") else ""
            print("  %-22s %-38s now:%-4s asked:%-9s%s"
                  % (row["protocol"], row["check"][:38], now, want or "-", flag))
            if row["note"]:
                print("      note: %s" % row["note"])

    # ---- what each answer is worth -----------------------------------------
    changed, unchanged, dropped = [], 0, []
    for row in answers:
        key = PROTOCOL_KEYS.get(str(row["protocol"]).strip())
        live = (cfg.get(key) or {}).get(str(row["check"]).strip()) if key else None
        if not live:
            problems.append("cannot find check %r under %r" % (row["check"], row["protocol"]))
            continue
        match = [o for o in live["options"] if o["label"].strip() == row["label"]]
        if len(match) != 1:
            problems.append("%s / %s: %d options match %r"
                            % (row["protocol"], row["check"], len(match), row["label"]))
            continue
        option = match[0]
        value = row["value"]
        if isinstance(value, tuple):
            problems.append("%s / %s / %s: cannot read %r as a score"
                            % (row["protocol"], row["check"], row["label"], value[1]))
            continue
        if value == "blank":
            if row["note"]:
                changed.append((row, live, option, "note only", None))
            continue
        if value == "drop":
            dropped.append((row, live, option))
            continue
        current = option.get("score", "missing")
        if current == value:
            unchanged += 1
            if row["note"]:
                changed.append((row, live, option, "note only", None))
            continue
        changed.append((row, live, option, current, value))

    print("\nSCORE CHANGES")
    if not changed and not dropped:
        print("  none -- every score in the workbook matches what ships")
    for row, live, option, current, value in changed:
        if current == "note only":
            print("  %-14s %-30s %-34s (score unchanged)"
                  % (row["protocol"], row["check"][:30], row["label"][:34]))
        else:
            print("  %-14s %-30s %-34s %s -> %s"
                  % (row["protocol"], row["check"][:30], row["label"][:34],
                     "none" if current is None else current,
                     "none" if value is None else value))
        if row["note"]:
            print("      note: %s" % row["note"])
    for row, live, option in dropped:
        print("  %-14s %-30s %-34s DROP THIS ANSWER"
              % (row["protocol"], row["check"][:30], row["label"][:34]))
        if row["note"]:
            print("      note: %s" % row["note"])

    print("\n  %d score change(s), %d answer(s) to drop, %d confirmed as-is"
          % (len([c for c in changed if c[3] != "note only"]), len(dropped), unchanged))

    if problems:
        print("\nCOULD NOT MATCH (%d) -- these need a human, nothing was applied for them:" % len(problems))
        for p in problems[:20]:
            print("  - %s" % p)

    if not args.apply:
        print("\nReport only. Re-run with --apply to write the score changes into plan-config.js.")
        return 0 if not problems else 1

    # ---- apply -------------------------------------------------------------
    # Matched on the option's LABEL within its own set, because that is what the
    # workbook shows and what the owner was looking at. A label that appears
    # twice in one set would be ambiguous -- the audit forbids duplicate codes,
    # and this refuses on duplicate labels for the same reason.
    if problems:
        print("\nRefusing to apply while %d row(s) could not be matched. Fix those first."
              % len(problems))
        return 1
    src = CONFIG.read_text(encoding="utf-8")
    applied = 0
    for row, live, option, current, value in changed:
        if current == "note only":
            continue
        label_sql = re.escape(option["label"])
        pattern = re.compile(
            r'(\{\s*code:\s*"%s",\s*label:\s*"%s",[^}]*?score:\s*)(null|\d+)'
            % (re.escape(option["code"]), label_sql))
        new_value = "null" if value is None else str(value)
        src, n = pattern.subn(lambda m: m.group(1) + new_value, src, count=1)
        if n != 1:
            print("  could not rewrite %s / %s -- left alone" % (row["check"], option["label"]))
            continue
        applied += 1
    CONFIG.write_text(src, encoding="utf-8")
    print("\nApplied %d score change(s) to %s" % (applied, CONFIG))
    if dropped:
        print("Answers marked DROP were NOT removed -- removing an answer changes what a "
              "technician can say, so do those by hand and re-run the audit.")
    print("Now run: node _qa/verify-check-controls.js && bash _qa/run_all.sh")
    return 0


if __name__ == "__main__":
    sys.exit(main())
