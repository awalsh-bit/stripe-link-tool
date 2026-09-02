#!/usr/bin/env python3
"""
Build the protocol review workbook -- the thing Cayden can actually edit.

The published review page is read-only HTML, and this session cannot read an
artifact back, so a page he edits in the browser could not reach me. A workbook
round-trips: he edits the yellow columns, sends the file back, and
apply-review-workbook.py writes his decisions into plan-config.js.

Three sheets, in the order the decisions happen:
  1. Start here      -- what to do, and what I am actually asking
  2. Should it score -- 76 checks: does this belong in the number at all
  3. What it's worth -- 149 answers: is my draft score right

Generated from check-table.json, so it cannot describe a protocol that is not
the one shipping. Run dump-check-table.js first.
"""
from __future__ import annotations

import json
import pathlib

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = pathlib.Path(__file__).parent
DATA = json.loads((HERE / "check-table.json").read_text(encoding="utf-8"))
OUT = HERE / "Wilson-Protocol-Review.xlsx"

FONT = "Arial"
GREEN = "1F6B39"
INK = "17251D"
MUTED = "63736A"

# Yellow is the fill this project uses for "you fill this in".
INPUT_FILL = PatternFill("solid", fgColor="FFF6D6")
HEAD_FILL = PatternFill("solid", fgColor="12331F")
BAND_FILL = PatternFill("solid", fgColor="EEF4EF")
NOTE_FILL = PatternFill("solid", fgColor="F7F9F7")
THIN = Side(style="thin", color="D7E2DA")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PROTOCOL_NAMES = {
    "generic": "Generic fallback", "refrigerator": "Refrigeration",
    "dishwasher": "Dishwasher", "cooking": "Range, cooktop & oven",
    "icemaker": "Icemaker (IMUC)", "washer": "Washer",
    "laundry": "Laundry centre", "dryer": "Dryer",
    "ventilation": "Vent hood", "microwave": "Microwave & speed oven",
    "outdoor_grill": "Outdoor grill", "hvac_cooling": "Cooling - split system",
    "hvac_heatpump": "Heat pump", "hvac_furnace": "Furnace",
    "hvac_minisplit": "Mini-split / ductless",
}
KIND_WORDS = {"scored": "Measured", "observed": "Condition", "trend": "Recorded only"}


EXAMPLE_MARK = "EXAMPLE ROW"


def example_row(ws, row, ncols, context, inputs):
    """One row showing the expected format, marked so the reader skips it.

    Kept OUT of the real data. A first draft put realistic values in the first
    genuine row -- refrigeration's frost pattern -- which meant an untouched
    workbook came back looking as though a decision had been made about it.
    """
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.value = context[i - 1] if i <= len(context) else inputs.get(i)
        c.fill = PatternFill("solid", fgColor="FBF3E6")
        c.font = Font(name=FONT, size=9, italic=True, color="8A4F1F")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 28
    return row + 1


def head(ws, row, labels, widths):
    for i, (label, width) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def body(cell, *, bold=False, wrap=True, fill=None, size=10, colour=INK, align="top"):
    cell.font = Font(name=FONT, size=size, bold=bold, color=colour)
    cell.alignment = Alignment(vertical=align, wrap_text=wrap)
    cell.border = BOX
    if fill:
        cell.fill = fill
    return cell


# ---------------------------------------------------------------------------
# 1. Start here
# ---------------------------------------------------------------------------
def sheet_start(wb, totals):
    ws = wb.create_sheet("Start here")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 104

    lines = [
        ("Wilson maintenance protocol - score review", "title"),
        ("", ""),
        ("Every question the field tool asks, and what every answer is worth. The scores are a "
         "draft written by someone who has never had to decide in a customer's kitchen whether a "
         "frost pattern is a problem. You have. Correct them.", "lede"),
        ("", ""),
        ("HOW TO USE THIS", "h"),
        ("Fill in the YELLOW columns only. Everything else is there for context and is safe to "
         "ignore. You do not have to answer every row - a sheet with fifteen corrections on it is "
         "worth more than a blank one.", "p"),
        ("Send the file back when you are done, or part-done. Nothing here has to be finished in "
         "one sitting.", "p"),
        ("", ""),
        ("THE TWO QUESTIONS", "h"),
        ("1.  'Should it score' sheet - does this check belong in the health score at all? Some "
         "checks are genuine indicators that something is failing. Some are just a record of what "
         "was seen. I have guessed; you know.", "p"),
        ("2.  'What it's worth' sheet - for the checks that do score, is the number behind each "
         "answer right? 5 of 5 means nothing wrong. 1 of 5 means failed. Blank means the answer "
         "scores nothing at all, which is what 'could not get to it' should always do.", "p"),
        ("", ""),
        ("THE RULES THESE NUMBERS FOLLOW", "h"),
        ("Dirt is never a deduction. Every 'cleaned at this visit' answer is a 5, the same as "
         "spotless. We do not dock a customer for the state their appliance was in before the "
         "technician arrived. If you see one of those scored below 5, that is a mistake - flag it.", "p"),
        ("'Could not get to it' scores nothing and is never held against the appliance. It is a "
         "better record than a guess.", "p"),
        ("Work we performed can never lift a score. That lives on the maintenance chips, not here.", "p"),
        ("A reading with no agreed target still does not score - oven accuracy, dryer exhaust, "
         "microwave heating, grill temperature. Those are recorded and compared against the same "
         "appliance's own history until there is a target worth judging them against.", "p"),
        ("", ""),
        ("WHAT IS IN THE PRODUCT TODAY", "h"),
        ("%d checks across %d protocols.  %d measured  ·  %d conditions judged by eye  ·  "
         "%d recorded only.  %d separate answers a technician can give."
         % (totals["checks"], totals["protocols"], totals["scored"], totals["observed"],
            totals["trend"], totals["answers"]), "p"),
        ("", ""),
        ("STILL OPEN - ANSWER ANYWHERE YOU LIKE", "h"),
        ("An undercounter cuber has a freeze plate, and scale or uneven freezing on it is the "
         "mechanism behind cloudy ice. There is no check for it. Should there be?", "p"),
        ("Anything the protocol does not ask that you would want on every stop?", "p"),
        ("Anything it asks that is not worth the minutes?", "p"),
    ]
    row = 2
    for text, kind in lines:
        c = ws.cell(row=row, column=2, value=text)
        if kind == "title":
            c.font = Font(name=FONT, size=16, bold=True, color=INK)
        elif kind == "lede":
            c.font = Font(name=FONT, size=11, color=INK)
            ws.row_dimensions[row].height = 46
        elif kind == "h":
            c.font = Font(name=FONT, size=10, bold=True, color=GREEN)
        else:
            c.font = Font(name=FONT, size=10, color="3C4A42")
            ws.row_dimensions[row].height = 14 + 12 * (len(text) // 100)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    key = row + 1
    ws.cell(row=key, column=2, value="KEY").font = Font(name=FONT, size=10, bold=True, color=GREEN)
    for offset, (label, fill) in enumerate([
        ("Yellow cells are yours to fill in.", INPUT_FILL),
        ("Everything else is context. Leave it alone.", NOTE_FILL),
    ], start=1):
        c = ws.cell(row=key + offset, column=2, value=label)
        c.font = Font(name=FONT, size=10, color="3C4A42")
        c.fill = fill
        c.border = BOX
    return ws


# ---------------------------------------------------------------------------
# 2. Should it score
# ---------------------------------------------------------------------------
def sheet_checks(wb, protocols):
    ws = wb.create_sheet("Should it score")
    head(ws,
         1,
         ["Appliance", "Check", "What the tech is asked to do", "How it is answered",
          "Counts toward the score today?", "SHOULD IT COUNT?", "YOUR NOTES"],
         [22, 30, 52, 22, 16, 16, 44])

    dv = DataValidation(type="list", formula1='"Yes,No,Not sure"', allow_blank=True)
    dv.error = "Choose Yes, No, or Not sure."
    dv.prompt = "Should this check move the customer's health score?"
    ws.add_data_validation(dv)

    row = example_row(ws, 2, 7,
                      ["EXAMPLE ROW - ignore or delete", "Ice pattern & production",
                       "(this row is here to show the format)", "Condition", "Yes"],
                      {6: "Yes", 7: "Say why, in your words. 'Cloudy ice is nearly always scale "
                                  "or a charge problem.'"})
    for proto in protocols:
        for check in proto["checks"]:
            band = BAND_FILL if (row % 2 == 0) else None
            body(ws.cell(row=row, column=1, value=PROTOCOL_NAMES.get(proto["key"], proto["key"])), fill=band)
            body(ws.cell(row=row, column=2, value=check["name"]), bold=True, fill=band)
            body(ws.cell(row=row, column=3, value=check["prompt"]), fill=band, size=9)
            body(ws.cell(row=row, column=4, value=KIND_WORDS.get(check["kind"], check["kind"])), fill=band, size=9)
            body(ws.cell(row=row, column=5, value="Yes" if check["scores"] else "No"),
                 fill=band, wrap=False,
                 colour=GREEN if check["scores"] else MUTED)
            answer = body(ws.cell(row=row, column=6), fill=INPUT_FILL, wrap=False)
            answer.alignment = Alignment(horizontal="center", vertical="top")
            dv.add(answer)
            body(ws.cell(row=row, column=7), fill=INPUT_FILL, size=9)
            ws.row_dimensions[row].height = 30
            row += 1
    return ws, row - 1


# ---------------------------------------------------------------------------
# 3. What it's worth
# ---------------------------------------------------------------------------
def sheet_answers(wb, protocols):
    ws = wb.create_sheet("What it's worth")
    head(ws,
         1,
         ["Appliance", "Check", "The answer a tech can tap", "What it means",
          "Draft score", "YOUR SCORE", "YOUR NOTES"],
         [20, 28, 46, 20, 11, 12, 40])

    dv = DataValidation(type="list", formula1='"5,4,3,2,1,none,drop"', allow_blank=True)
    dv.error = "Use 5 4 3 2 1, or 'none' for scores nothing, or 'drop' to remove the answer."
    dv.prompt = ("5 = nothing wrong. 1 = failed. 'none' = scores nothing at all. "
                 "'drop' = this answer should not exist.")
    ws.add_data_validation(dv)

    row = example_row(ws, 2, 7,
                      ["EXAMPLE ROW - ignore or delete", "Ice pattern & production",
                       "Cloudy or incomplete cubes", "Abnormal", "3 of 5"],
                      {6: "2", 7: "Cloudy ice is worse than a 3 - it is nearly always scale."})
    count = 0
    for proto in protocols:
        for check in proto["checks"]:
            if not check["options"]:
                continue
            for i, opt in enumerate(check["options"]):
                band = BAND_FILL if (row % 2 == 0) else None
                body(ws.cell(row=row, column=1,
                             value=PROTOCOL_NAMES.get(proto["key"], proto["key"]) if i == 0 else ""),
                     fill=band, size=9)
                body(ws.cell(row=row, column=2, value=check["name"] if i == 0 else ""),
                     bold=(i == 0), fill=band, size=9)
                body(ws.cell(row=row, column=3, value=opt["label"]), fill=band)
                body(ws.cell(row=row, column=4, value=opt["result"]), fill=band, size=9, colour=MUTED)
                draft = opt["rating"]
                shown = "" if not check["scores"] else ("none" if draft is None else "%d of 5" % draft)
                cell = body(ws.cell(row=row, column=5, value=shown), fill=band, wrap=False)
                cell.alignment = Alignment(horizontal="center", vertical="top")
                if draft is not None and check["scores"]:
                    cell.font = Font(name=FONT, size=10,
                                     bold=draft <= 2,
                                     color=GREEN if draft == 5 else ("8A4A17" if draft <= 2 else "7D5512"))
                mine = body(ws.cell(row=row, column=6), fill=INPUT_FILL, wrap=False)
                mine.alignment = Alignment(horizontal="center", vertical="top")
                dv.add(mine)
                body(ws.cell(row=row, column=7), fill=INPUT_FILL, size=9)

                if opt["needs"] == "code":
                    ws.cell(row=row, column=3).comment = Comment(
                        "Choosing this makes the tech record the actual code before the check "
                        "can be completed.", "Wilson", width=260, height=70)
                if not check["scores"]:
                    ws.cell(row=row, column=5).value = "n/a"
                    ws.cell(row=row, column=5).font = Font(name=FONT, size=9, color=MUTED)
                    ws.cell(row=row, column=5).comment = Comment(
                        "This check does not score at all today. If it should, say so on the "
                        "'Should it score' sheet and put a number here.", "Wilson",
                        width=280, height=80)
                ws.row_dimensions[row].height = 26
                row += 1
                count += 1
    return ws, count


def build():
    protocols = ([p for p in DATA["protocols"] if not p["isHvac"] and p["key"] != "generic"]
                 + [p for p in DATA["protocols"] if p["isHvac"]]
                 + [p for p in DATA["protocols"] if p["key"] == "generic"])

    answers = sum(len(c["options"]) for p in protocols for c in p["checks"])
    totals = dict(DATA["totals"])
    totals["answers"] = answers

    wb = Workbook()
    wb.remove(wb.active)
    sheet_start(wb, totals)
    _, n_checks = sheet_checks(wb, protocols)
    _, n_answers = sheet_answers(wb, protocols)
    wb.save(OUT)
    print("wrote %s -- %d checks, %d answers" % (OUT, n_checks, n_answers))


if __name__ == "__main__":
    build()
